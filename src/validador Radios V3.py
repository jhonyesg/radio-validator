import subprocess
import pandas as pd
import os
import tkinter as tk
from tkinter import ttk, filedialog
from tkinter import messagebox
import configparser
import pyperclip
import json
import threading
from ttkthemes import ThemedTk
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import requests
import vlc
import csv
import getpass

# Obtener el directorio del script actual
script_dir = os.path.dirname(os.path.abspath(__file__))

# Variables globales para el reproductor VLC
playing = False
player = None
current_stream = None
all_items = []
csv_imported = False
csv_data = []

# Variable global para controlar la validación
validacion_en_curso = False

# Obtener el nombre del usuario actual
current_user = getpass.getuser()

# Lista de rutas predefinidas de archivos .ini (solo el del proyecto)
predefined_ini_paths = [
    os.path.join(script_dir, "..", "radios_part1.ini")
]


def filter_treeview(event=None):
    search_term = search_var.get().lower()
    for item in tree.get_children():
        tree.delete(item)

    for item in all_items:
        values = item['values']
        if not search_term or search_term in str(values).lower():
            tree.insert('', 'end', values=values, tags=item['tags'])


def clear_filter():
    filter_entry.delete(0, tk.END)
    for item in tree.get_children():
        tree.delete(item)
    for item in all_items:
        values = item['values']
        tree.insert('', 'end', values=values, tags=item['tags'])


def copy_url():
    selected_items = tree.selection()
    if not selected_items:
        messagebox.showinfo(
            "Información", "Por favor, selecciona un elemento de la tabla primero.")
        return
    url = tree.item(selected_items[0])['values'][1]
    pyperclip.copy(url)
    messagebox.showinfo(
        "URL Copiada", "El URL del stream ha sido copiado al portapapeles.")


def play_pause():
    global playing, player, current_stream
    selected_items = tree.selection()
    if not selected_items:
        messagebox.showinfo(
            "Información", "Por favor, selecciona un elemento de la tabla primero.")
        return
    url = tree.item(selected_items[0])['values'][1]
    if not playing:
        if current_stream != url:
            if player:
                player.stop()
            player = vlc.MediaPlayer(url)
            current_stream = url
        if player is not None:
            player.play()
            playing = True
            vlc_button.config(text="Pausar")
        else:
            messagebox.showerror("Error", "No se pudo inicializar el reproductor VLC para este enlace.")
    else:
        if player is not None:
            player.pause()
        playing = False
        vlc_button.config(text="Reproducir")


def stop():
    global playing, player
    if player:
        player.stop()
        playing = False
        vlc_button.config(text="Reproducir")


def set_volume(val):
    try:
        # Primero convierte la cadena a float, luego a int
        volume = int(float(val))

        if player is None:
            print("Error: El reproductor no está inicializado.")
            return

        if not hasattr(player, 'audio_set_volume'):
            print("Error: El reproductor no tiene el método 'audio_set_volume'.")
            return

        player.audio_set_volume(volume)
        print(f"Volumen establecido a {volume}")
    except ValueError as e:
        print(f"Error al convertir el valor del volumen: {e}")
    except Exception as e:
        print(f"Error al establecer el volumen: {e}")


def select_config_file():
    file_path = filedialog.askopenfilename(
        filetypes=[("INI files", "*.ini"), ("CSV files", "*.csv")])
    if file_path:
        config_entry.delete(0, tk.END)
        config_entry.insert(0, file_path)


def update_config_entry(*args):
    selected_path = config_var.get()
    config_entry.delete(0, tk.END)
    config_entry.insert(0, selected_path)


def import_csv():
    global csv_imported, csv_data
    file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    if file_path:
        csv_data = []
        with open(file_path, 'r') as csvfile:
            csvreader = csv.DictReader(csvfile)
            for row in csvreader:
                csv_data.append(row)

        # Limpiar la tabla actual
        for item in tree.get_children():
            tree.delete(item)

        # Limpiar all_items
        all_items.clear()

        # Agregar los datos del CSV a la tabla y a all_items
        for row in csv_data:
            medio = row['source'].strip('"')
            url = row['target']
            output_dir = ""  # No hay un campo equivalente en el CSV
            status = row['status']
            info = f"Type: {row['type']}, Hits: {row['hits']}, Title: {row['title']}"

            tag = 'operativo' if status == 'active' else 'caido'

            tree.insert("", tk.END, values=(
                medio, url, output_dir, status, info), tags=(tag,))
            all_items.append({
                'values': (medio, url, output_dir, status, info),
                'tags': (tag,)
            })

        csv_imported = True
        config_entry.delete(0, tk.END)
        config_entry.insert(0, file_path)
        messagebox.showinfo(
            "CSV Importado", f"Se ha importado el archivo CSV: {file_path}")


# Crear una ventana principal con tema
window = ThemedTk(theme="default")
window.title("Validación de Enlaces de Streaming")
window.geometry("1200x800")

# Configurar el ícono de la ventana
icon_path = os.path.join(script_dir, 'logo.ico')
try:
    window.iconbitmap(icon_path)
except tk.TclError:
    print(
        f"No se pudo cargar el ícono '{icon_path}'. Asegúrate de que el archivo existe en el directorio correcto.")

# Configurar colores personalizados
bg_color = "#ffffff"
fg_color = "#000000"
title_color = "#000000"
vlc_button_color = "#ff8800"
copy_button_color = "#3584e4"
validate_button_color = "#cc0000"
tree_bg = "#ffffff"
tree_fg = "#000000"
green_button_color = "#4CAF50"

window.configure(bg=bg_color)

# Configurar estilos
style = ttk.Style()
style.theme_use('default')
style.configure(".", background=bg_color, foreground=fg_color)
style.configure("TFrame", background=bg_color)
style.configure("TLabel", background=bg_color, foreground=fg_color)
style.configure("TEntry", fieldbackground=bg_color, foreground=fg_color)
style.configure("Treeview", background=tree_bg,
                foreground=tree_fg, fieldbackground=tree_bg)
style.map("Treeview", background=[('selected', copy_button_color)])
style.configure("TProgressbar", background=copy_button_color)

# Configurar estilos de botones
style.configure("VLC.TButton", background=vlc_button_color,
                foreground="#ffffff", padding=10)
style.configure("Copy.TButton", background=copy_button_color,
                foreground="#ffffff", padding=10)
style.configure("Validate.TButton", background=validate_button_color,
                foreground="#ffffff", padding=10)

style.configure("Green.TButton", background=green_button_color,
                foreground="#ffffff", padding=10)

style.map("VLC.TButton", background=[('active', vlc_button_color)])
style.map("Copy.TButton", background=[('active', copy_button_color)])
style.map("Validate.TButton", background=[('active', validate_button_color)])
style.map("Green.TButton", background=[('active', green_button_color)])

# Título centrado
title_label = ttk.Label(window, text="Validación de Enlaces de Streaming", font=(
    "Helvetica", 16, "bold"), foreground=title_color)
title_label.pack(pady=10)

# Frame principal
main_frame = ttk.Frame(window)
main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

# Frame para seleccionar archivo de configuración
config_frame = ttk.Frame(main_frame)
config_frame.pack(fill=tk.X, pady=5)

config_label = ttk.Label(config_frame, text="Archivo de configuración:")
config_label.pack(side=tk.LEFT, padx=(0, 5))

config_var = tk.StringVar()
config_dropdown = ttk.Combobox(
    config_frame, textvariable=config_var, values=predefined_ini_paths, state="readonly")
config_dropdown.pack(side=tk.LEFT, expand=True, fill=tk.X)
config_dropdown.set(predefined_ini_paths[0] if predefined_ini_paths else "")
config_var.trace('w', update_config_entry)

config_entry = ttk.Entry(config_frame)
config_entry.pack(side=tk.LEFT, expand=True, fill=tk.X)
config_entry.insert(0, predefined_ini_paths[0] if predefined_ini_paths else "")

config_button = ttk.Button(
    config_frame, text="Seleccionar", command=select_config_file)
config_button.pack(side=tk.LEFT, padx=5)

import_csv_button = ttk.Button(
    config_frame, text="Importar CSV", command=import_csv)
import_csv_button.pack(side=tk.LEFT, padx=5)

# Crear una etiqueta para mostrar el estado de validación
status_label = ttk.Label(main_frame, text="Estado de validación:")
status_label.pack(pady=5)

# Crear un campo de entrada para el filtro
filter_frame = ttk.Frame(main_frame)
filter_frame.pack(fill=tk.X, pady=5)

filter_label = ttk.Label(filter_frame, text="Filtrar:")
filter_label.pack(side=tk.LEFT, padx=(0, 5))

search_var = tk.StringVar()
filter_entry = ttk.Entry(filter_frame, textvariable=search_var)
filter_entry.pack(side=tk.LEFT, expand=True, fill=tk.X)
filter_entry.bind('<KeyRelease>', filter_treeview)

clear_filter_button = ttk.Button(
    filter_frame, text="Limpiar Filtro", command=clear_filter)
clear_filter_button.pack(side=tk.LEFT, padx=5)

# Frame para la tabla y la barra de desplazamiento
tree_frame = ttk.Frame(main_frame)
tree_frame.pack(fill=tk.BOTH, expand=True)

# Crear un Treeview para mostrar los resultados en forma de tabla
tree = ttk.Treeview(tree_frame, columns=(
    "Medio", "Streaming URL", "Output Directory", "Status", "Info"), show="headings")
tree.heading("Medio", text="Medio")
tree.heading("Streaming URL", text="Streaming URL")
tree.heading("Output Directory", text="Output Directory")
tree.heading("Status", text="Status")
tree.heading("Info", text="Info")

# Ajustar el ancho de las columnas
tree.column("Medio", width=150)
tree.column("Streaming URL", width=300)
tree.column("Output Directory", width=300)
tree.column("Status", width=100)
tree.column("Info", width=250)

# Añadir barra de desplazamiento
scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
tree.configure(yscrollcommand=scrollbar.set)

# Empaquetar el árbol y la barra de desplazamiento
tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Configurar los colores de fondo
tree.tag_configure('caido', background='#ffcccc')
tree.tag_configure('operativo', background='#ccffcc')

# Frame para los botones y la lista desplegable
button_frame = ttk.Frame(main_frame)
button_frame.pack(fill=tk.X, pady=10)

# Botones para copiar URL y reproducir/pausar
copy_button = ttk.Button(button_frame, text="Copiar Enlace",
                         command=copy_url, style="Copy.TButton")
copy_button.pack(side=tk.LEFT, padx=5)

vlc_button = ttk.Button(button_frame, text="Reproducir",
                        command=play_pause, style="VLC.TButton")
vlc_button.pack(side=tk.LEFT, padx=5)

stop_button = ttk.Button(button_frame, text="Detener",
                         command=stop, style="VLC.TButton")
stop_button.pack(side=tk.LEFT, padx=5)

# Control de volumen
volume_label = ttk.Label(button_frame, text="Volumen:")
volume_label.pack(side=tk.LEFT, padx=5)

volume_scale = ttk.Scale(button_frame, from_=0, to=100,
                         orient=tk.HORIZONTAL, command=set_volume)
volume_scale.set(100)
volume_scale.pack(side=tk.LEFT, padx=5)

# Lista desplegable para seleccionar la acción
action_var = tk.StringVar()
action_combobox = ttk.Combobox(button_frame, textvariable=action_var,
                               values=["Iniciar Validación",
                                       "Exportar a Excel"],
                               state="readonly")
action_combobox.pack(side=tk.LEFT, padx=5)
action_combobox.set("Iniciar Validación")

# Barra de progreso
progress_bar = ttk.Progressbar(
    main_frame, orient=tk.HORIZONTAL, length=300, mode='determinate')
progress_bar.pack(pady=10)


def check_stream_with_ffprobe(url):
    try:
        result = subprocess.run(["ffprobe", "-v", "quiet", "-print_format", "json", "-show_format", "-show_streams", url],
                                capture_output=True, text=True, timeout=10)
        if result.returncode == 0:
            data = json.loads(result.stdout)
            if 'streams' in data and len(data['streams']) > 0:
                return True, f"Codec: {data['streams'][0].get('codec_name', 'N/A')}, Bitrate: {data['format'].get('bit_rate', 'N/A')}"
        return False, "No se pudo obtener información del stream"
    except subprocess.TimeoutExpired:
        return False, "Timeout al obtener información"
    except Exception as e:
        return False, f"Error: {str(e)}"


def validate_link(url):
    try:
        response = requests.head(url, timeout=10)
        return response.status_code == 200
    except:
        return False


def validar_enlaces():
    global all_items, validacion_en_curso, csv_imported, csv_data
    validacion_en_curso = True
    config_file = config_entry.get()
    if not config_file:
        messagebox.showerror(
            "Error", "Por favor, seleccione un archivo de configuración o importe un CSV.")
        validacion_en_curso = False
        return

    # Limpiar la tabla antes de iniciar una nueva validación
    for i in tree.get_children():
        tree.delete(i)

    # Limpiar la lista all_items
    all_items = []

    # Determinar si es un archivo INI, CSV importado, o CSV seleccionado
    if csv_imported:
        streaming_urls = [row['target'] for row in csv_data]
        output_dirs = ["" for _ in csv_data]  # No hay campo equivalente en CSV
        medios = [row['source'].strip('"') for row in csv_data]
    elif config_file.lower().endswith('.ini'):
        # Leer el archivo INI seleccionado
        config = configparser.ConfigParser()
        config.read(config_file)

        # Extraer los datos necesarios del archivo INI
        streaming_urls = []
        output_dirs = []
        medios = []

        for section in config.sections():
            streaming_urls.append(config[section]['stream_url'])
            output_dirs.append(config[section]['output_dir'])
            medios.append(section)
    elif config_file.lower().endswith('.csv'):
        # Leer el archivo CSV seleccionado
        with open(config_file, 'r') as csvfile:
            csvreader = csv.DictReader(csvfile)
            streaming_urls = []
            output_dirs = []
            medios = []
            for row in csvreader:
                medios.append(row['source'].strip('"'))
                streaming_urls.append(row['target'])
                output_dirs.append("")  # No hay un campo equivalente en el CSV
    else:
        messagebox.showerror(
            "Error", "Formato de archivo no soportado. Por favor, seleccione un archivo .ini o .csv.")
        validacion_en_curso = False
        return

    # Configurar la barra de progreso
    progress_bar['maximum'] = len(streaming_urls)
    progress_bar['value'] = 0

    # Iterar sobre los enlaces
    for i, url in enumerate(streaming_urls):
        if not validacion_en_curso:
            break

        # Actualizar la etiqueta de estado de validación
        status_label.config(text=f"Validando {medios[i]}: {url}")
        window.update()

        # Primero, intentar con requests
        if validate_link(url):
            status = "Enlace Operativo"
            tag = 'operativo'
            info = "Validado con requests"
        else:
            # Si requests falla, intentar con ffprobe
            ffprobe_status, ffprobe_info = check_stream_with_ffprobe(url)

            if ffprobe_status:
                status = "Enlace Operativo"
                tag = 'operativo'
                info = ffprobe_info
            else:
                # Si ffprobe falla, intentar con ffmpeg
                try:
                    process = subprocess.run(["ffmpeg", "-i", url, "-t", "2", "-f", "mp3", "-y", "temp.mp3"],
                                             capture_output=True, timeout=10)
                    if process.returncode == 0 and os.path.exists("temp.mp3") and os.stat("temp.mp3").st_size > 0:
                        status = "Enlace Operativo"
                        tag = 'operativo'
                        info = "Validado con ffmpeg"
                    else:
                        status = "Enlace Caido"
                        tag = 'caido'
                        info = "Fallo en requests, ffprobe y ffmpeg"
                except subprocess.TimeoutExpired:
                    status = "Enlace Caido"
                    tag = 'caido'
                    info = "Timeout en requests, ffprobe y ffmpeg"

        # Agregar el resultado a la tabla
        item = tree.insert("", tk.END, values=(
            medios[i], url, output_dirs[i], status, info), tags=(tag,))
        all_items.append({
            'values': (medios[i], url, output_dirs[i], status, info),
            'tags': (tag,)
        })

        # Actualizar la barra de progreso
        progress_bar['value'] = i + 1
        window.update()

    # Borrar archivo temporal si existe
    if os.path.exists("temp.mp3"):
        os.remove("temp.mp3")

    # Mostrar un mensaje de finalización
    if validacion_en_curso:
        messagebox.showinfo("Validación Completada",
                            "La validación de enlaces ha finalizado.")
    else:
        messagebox.showinfo("Validación Interrumpida",
                            "La validación de enlaces ha sido interrumpida.")

    # Resetear la barra de progreso
    progress_bar['value'] = 0
    validacion_en_curso = False
    csv_imported = False


def detener_validacion():
    global validacion_en_curso
    validacion_en_curso = False
    messagebox.showinfo("Deteniendo", "Deteniendo la validación...")


def export_to_excel():
    # Crear un nuevo libro de trabajo de Excel
    wb = Workbook()
    if wb.active is not None:
        ws = wb.active
    else:
        ws = wb.create_sheet(title="Validación de Enlaces")
    ws.title = "Validación de Enlaces"

    # Definir los estilos
    header_fill = PatternFill(start_color="3584E4",
                              end_color="3584E4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    operativo_fill = PatternFill(
        start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    caido_fill = PatternFill(start_color="FFCCCC",
                             end_color="FFCCCC", fill_type="solid")

    # Escribir los encabezados
    headers = ["Medio", "Streaming URL", "Output Directory", "Status", "Info"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Escribir los datos
    for row, item in enumerate(tree.get_children(), start=2):
        values = tree.item(item)['values']
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if values[3] == "Enlace Operativo":
                cell.fill = operativo_fill
            elif values[3] == "Enlace Caido":
                cell.fill = caido_fill

    # Ajustar el ancho de las columnas
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 30

    # Guardar el archivo
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        wb.save(file_path)
        messagebox.showinfo("Exportación Completada",
                            f"Los datos han sido exportados a {file_path}")


def perform_action():
    action = action_var.get()
    if action == "Iniciar Validación":
        threading.Thread(target=validar_enlaces, daemon=True).start()
    elif action == "Exportar a Excel":
        export_to_excel()


# Botón para realizar la acción seleccionada
action_button = ttk.Button(button_frame, text="Realizar Acción",
                           command=perform_action, style="Green.TButton")
action_button.pack(side=tk.LEFT, padx=5)

# Botón para detener la validación
stop_validation_button = ttk.Button(button_frame, text="Detener Validación",
                                    command=detener_validacion, style="Validate.TButton")
stop_validation_button.pack(side=tk.LEFT, padx=5)

# Ejecutar el bucle de eventos de la ventana principal
window.mainloop()
